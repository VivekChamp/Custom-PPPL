
import frappe
from frappe.desk.reportview import clean_params, parse_json
from frappe import _
import json
from frappe.desk.query_report import (
	run
)
import datetime


from frappe.utils import cint, cstr,format_duration

@frappe.whitelist()
def export_query():
	"""export from query reports"""
	from frappe.desk.utils import get_csv_bytes, pop_csv_params, provide_binary_file

	form_params = frappe._dict(frappe.local.form_dict)
	csv_params = pop_csv_params(form_params)
	clean_params(form_params)
	parse_json(form_params)

	report_name = form_params.report_name
	frappe.permissions.can_export(
		frappe.get_cached_value("Report", report_name, "ref_doctype"),
		raise_exception=True,
	)

	file_format_type = form_params.file_format_type
	custom_columns = frappe.parse_json(form_params.custom_columns or "[]")
	include_indentation = form_params.include_indentation
	include_filters = form_params.include_filters
	visible_idx = form_params.visible_idx

	if isinstance(visible_idx, str):
		visible_idx = json.loads(visible_idx)

	data = run(report_name, form_params.filters, custom_columns=custom_columns, are_default_filters=False)
	data = frappe._dict(data)
	data.filters = form_params.applied_filters

	if not data.columns:
		frappe.respond_as_web_page(
			_("No data to export"),
			_("You can try changing the filters of your report."),
		)
		return

	format_duration_fields(data)
	xlsx_data, column_widths = build_xlsx_data(
		data, visible_idx, include_indentation, report_name, include_filters=include_filters
	)
 
	if file_format_type == "CSV":
		content = get_csv_bytes(xlsx_data, csv_params)
		file_extension = "csv"
	elif file_format_type == "Excel":
		from frappe.utils.xlsxutils import make_xlsx

		file_extension = "xlsx"
  
		content = write_wrapped_xlsx((xlsx_data, column_widths), report_name)
  
	provide_binary_file(report_name, file_extension, content)


def format_duration_fields(data: frappe._dict) -> None:
	for i, col in enumerate(data.columns):
		if col.get("fieldtype") != "Duration":
			continue

		for row in data.result:
			index = col.get("fieldname") if isinstance(row, dict) else i
			if row[index]:
				row[index] = format_duration(row[index])


def build_xlsx_data(data, visible_idx, include_indentation, report_name, include_filters=False, ignore_visible_idx=False):
	EXCEL_TYPES = (
		str,
		bool,
		type(None),
		int,
		float,
		datetime.datetime,
		datetime.date,
		datetime.time,
		datetime.timedelta,
	)

	if len(visible_idx) == len(data.result) or not visible_idx:
		# It's not possible to have same length and different content.
		ignore_visible_idx = True
	else:
		# Note: converted for faster lookups
		visible_idx = set(visible_idx)

	result = []
	column_widths = []

	if cint(include_filters):
		filter_data = []
		branch_name = []
		customer_list = []
		supplier_list = []
		party_type = None
		party_value = None
  
		filters = data.filters
		for filter_name, filter_value in filters.items():
			if not filter_value:
				continue
			filter_value = (
				", ".join([cstr(x) for x in filter_value])
				if isinstance(filter_value, list)
				else cstr(filter_value)
			)
			filter_data.append([cstr(filter_name), filter_value])
			
			if report_name == "Custom Accounts Receivable" or report_name == "Custom General Ledger":
				if filter_name.lower() == "branch":
					# Split each element by comma and flatten into one list
					if isinstance(filter_value, list):
						branch_name = []
						for b in filter_value:
							branch_name += [x.strip() for x in b.split(",")]
					else:
						branch_name = [x.strip() for x in filter_value.split(",")]
      

				for f in filter_data:
					label = f[0].strip().lower()
					value = f[1].strip()

					if label == "party type":
						party_type = value
					elif label == "party":
						party_value = value

		
		if party_value:
			party_names = [p.strip() for p in party_value.split(",")]

			if party_type == "Customer":
				customer_list = party_names
			elif party_type == "Supplier":
				supplier_list = party_names


		filter_data.append([])
    
		if branch_name:
			for branch in branch_name:
				
				branch_address_name = frappe.db.get_value("Branch", branch, "custom_address")
				if branch_address_name:
					doc = frappe.get_doc("Address", branch_address_name)
					branch_address = (
						f"{branch}\n"
						f"{doc.address_line1 or ''}\n"
						f"{doc.city or ''}, {doc.state or ''} - {doc.pincode or ''}"
					)
					filter_data.append([("Branch Address"), branch_address])
		
		if customer_list:
			for customer in customer_list:
				
				primary_address_name = frappe.db.get_value("Customer", customer, "customer_primary_address")

				if not primary_address_name:
					
					primary_address_name = frappe.db.get_value(
						"Dynamic Link",
						filters={
							"link_doctype": "Customer",
							"link_name": customer,
							"parenttype": "Address"
						},
						fieldname="parent"
					)

				if primary_address_name:
					address_doc = frappe.get_doc("Address", primary_address_name)
					address_plain = f"{address_doc.address_line1 or ''}\n{address_doc.city or ''}, {address_doc.state or ''} - {address_doc.pincode or ''}"

					
					filter_data.append(["Customer", f"{customer}\n{address_plain}"])
     
		if supplier_list:
			for supplier in supplier_list:
				
				primary_address_name = frappe.db.get_value("Supplier", supplier, "supplier_primary_address")

				if not primary_address_name:
					
					primary_address_name = frappe.db.get_value(
						"Dynamic Link",
						filters={
							"link_doctype": "Supplier",
							"link_name": supplier,
							"parenttype": "Address"
						},
						fieldname="parent"
					)

				if primary_address_name:
					address_doc = frappe.get_doc("Address", primary_address_name)
					address_plain = f"{address_doc.address_line1 or ''}\n{address_doc.city or ''}, {address_doc.state or ''} - {address_doc.pincode or ''}"

					
					filter_data.append(["Supplier", f"{supplier}\n{address_plain}"])

		result += filter_data

	column_data = []
	for column in data.columns:
		if column.get("hidden"):
			continue
		column_data.append(_(column.get("label")))
		column_width = cint(column.get("width", 0))
		# to convert into scale accepted by openpyxl
		column_width /= 10
		column_widths.append(column_width)
	result.append(column_data)

	# build table from result
	for row_idx, row in enumerate(data.result):
		# only pick up rows that are visible in the report
		if ignore_visible_idx or row_idx in visible_idx:
			row_data = []
			if isinstance(row, dict):
				for col_idx, column in enumerate(data.columns):
					if column.get("hidden"):
						continue
					label = column.get("label")
					fieldname = column.get("fieldname")
					cell_value = row.get(fieldname, row.get(label, ""))
					if not isinstance(cell_value, EXCEL_TYPES):
						cell_value = cstr(cell_value)

					if cint(include_indentation) and "indent" in row and col_idx == 0:
						cell_value = ("    " * cint(row["indent"])) + cstr(cell_value)
					row_data.append(cell_value)
			elif row:
				row_data = row

			result.append(row_data)
	for col_idx in range(len(column_widths)):
		max_len = 0
		for row in result[1:]:  # skip header
			if len(row) > col_idx:
				cell = row[col_idx]
				if isinstance(cell, str):
					cell_lines = cell.split("\n")  # account for newlines
					line_lengths = [len(line) for line in cell_lines]
					max_len = max(max_len, *line_lengths)
				else:
					max_len = max(max_len, len(cstr(cell)))

		# Adjust scaling: Excel column width ~= string length * 1.2
		scaled_width = round(max_len * 1.2, 1)
		column_widths[col_idx] = max(column_widths[col_idx], scaled_width)
	return result, column_widths






def write_wrapped_xlsx(data, sheet_name="Report"):
	from io import BytesIO
	import openpyxl
	from openpyxl.styles import Alignment
	from openpyxl.utils import get_column_letter

	output = BytesIO()
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = sheet_name

	result, column_widths = data

	for row_idx, row in enumerate(result, 1):
		for col_idx, val in enumerate(row, 1):
			cell = ws.cell(row=row_idx, column=col_idx, value=val)

			# Wrap text if value contains line break
			if isinstance(val, str) and "\n" in val:
				cell.alignment = Alignment(wrap_text=True)
				ws.row_dimensions[row_idx].height = 40

	# Set column widths (optional)
	for idx, width in enumerate(column_widths, 1):
		col_letter = get_column_letter(idx)
		ws.column_dimensions[col_letter].width = max(width, 15)

	wb.save(output)
	output.seek(0)
	return output.getvalue()  # ✅ This is the key fix




